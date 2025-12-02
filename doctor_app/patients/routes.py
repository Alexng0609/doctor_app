from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required
from doctor_app import db
from doctor_app.models import Patient, Visit, Diagnosis
from doctor_app.patients.forms import PatientForm, ImportPatientsForm
import openpyxl
from datetime import datetime

bp = Blueprint("patients", __name__)


def check_duplicate_patient(full_name, phone, doctor_id, exclude_patient_id=None):
    """
    Check if a patient with the same name and phone already exists for this doctor.

    Args:
        full_name: Patient's full name
        phone: Patient's phone number
        doctor_id: Doctor's user ID
        exclude_patient_id: Patient ID to exclude from check (for editing)

    Returns:
        Existing patient if duplicate found, None otherwise
    """
    # Normalize inputs
    full_name = full_name.strip() if full_name else ""
    phone = phone.strip() if phone else ""

    # Build query
    query = Patient.query.filter_by(doctor_id=doctor_id)

    # Exclude current patient if editing
    if exclude_patient_id:
        query = query.filter(Patient.id != exclude_patient_id)

    # Strategy 1: Check by name AND phone (most reliable)
    if phone:
        duplicate = query.filter_by(full_name=full_name, phone=phone).first()
        if duplicate:
            return duplicate

    # Strategy 2: Check by name only (if no phone provided or no match with phone)
    # This catches cases where phone might be updated/changed
    duplicate = query.filter_by(full_name=full_name).first()
    if duplicate:
        # If we found a name match, check if phones are similar
        # Allow if one has no phone and other does (phone was added later)
        if not duplicate.phone or not phone:
            return duplicate
        # If both have phones but they're different, this might be a different person
        # with same name - don't block, but warn
        return None

    return None


# List all patients
@bp.route("/", methods=["GET"])
@login_required
def list_patients():
    from flask_login import current_user

    q = request.args.get("q", "")

    if q:
        # When searching, show ALL patients from all locations
        patients = (
            Patient.query.filter(
                (Patient.full_name.ilike(f"%{q}%")) | (Patient.phone.ilike(f"%{q}%"))
            )
            .order_by(Patient.full_name.asc())
            .all()
        )
    else:
        # Use permission-based query
        patients = (
            current_user.get_accessible_patients()
            .order_by(Patient.full_name.asc())
            .all()
        )

    return render_template(
        "patients/list.html", patients=patients, q=q, now=datetime.now()
    )


# Add new patient
@bp.route("/new", methods=["GET", "POST"])
@login_required
def new_patient():
    from flask_login import current_user

    form = PatientForm()
    if form.validate_on_submit():
        # Check for duplicates BEFORE creating
        duplicate = check_duplicate_patient(
            full_name=form.full_name.data,
            phone=form.phone.data,
            doctor_id=current_user.id,
        )

        if duplicate:
            # Found a duplicate - show error and link to existing patient
            flash(
                f"⚠️ Patient already exists! "
                f"'{duplicate.full_name}' with phone '{duplicate.phone or 'N/A'}' "
                f"is already in your patient list.",
                "danger",
            )
            flash(
                f"Click here to view existing patient: "
                f"<a href='{url_for('patients.patient_detail', patient_id=duplicate.id)}' style='color: white; text-decoration: underline;'>"
                f"View {duplicate.full_name}</a>",
                "warning",
            )
            # Re-render form with submitted data
            return render_template("patients/new.html", form=form)

        # No duplicate - create new patient
        p = Patient(
            full_name=form.full_name.data.strip(),
            phone=form.phone.data.strip() if form.phone.data else None,
            date_of_birth=form.date_of_birth.data,
            doctor_id=current_user.id,  # Assign to current doctor
        )
        db.session.add(p)
        db.session.commit()
        flash("✅ Patient created successfully. You can now add a visit.", "success")
        return redirect(url_for("patients.patient_detail", patient_id=p.id))

    return render_template("patients/new.html", form=form)


# Patient detail page
@bp.route("/<int:patient_id>", methods=["GET"])
@login_required
def patient_detail(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    visits = patient.visits
    return render_template(
        "patients/detail.html", patient=patient, visits=visits, now=datetime.now()
    )


# Edit patient
@bp.route("/<int:patient_id>/edit", methods=["GET", "POST"])
@login_required
def edit_patient(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    form = PatientForm(obj=patient)

    if form.validate_on_submit():
        # Check for duplicates BEFORE updating (exclude current patient)
        duplicate = check_duplicate_patient(
            full_name=form.full_name.data,
            phone=form.phone.data,
            doctor_id=patient.doctor_id,
            exclude_patient_id=patient.id,  # Exclude current patient from check
        )

        if duplicate:
            # Found a duplicate - show error
            flash(
                f"⚠️ Cannot update: Another patient with the same name and phone already exists! "
                f"'{duplicate.full_name}' with phone '{duplicate.phone or 'N/A'}' "
                f"(Patient ID: {duplicate.id})",
                "danger",
            )
            # Re-render form with submitted data
            return render_template("patients/edit.html", form=form, patient=patient)

        # No duplicate - update patient
        patient.full_name = form.full_name.data.strip()
        patient.phone = form.phone.data.strip() if form.phone.data else None
        patient.date_of_birth = form.date_of_birth.data
        db.session.commit()
        flash("✅ Patient updated successfully", "success")
        return redirect(url_for("patients.patient_detail", patient_id=patient.id))

    return render_template("patients/edit.html", form=form, patient=patient)


# Delete patient
@bp.route("/<int:patient_id>/delete", methods=["POST"])
@login_required
def delete_patient(patient_id):
    from flask_login import current_user

    # Permission check
    if not current_user.can_delete_patient():
        flash(
            "⛔ You don't have permission to delete patients. Contact your doctor.",
            "danger",
        )
        return redirect(url_for("patients.list_patients"))

    patient = Patient.query.get_or_404(patient_id)
    patient_name = patient.full_name

    db.session.delete(patient)
    db.session.commit()
    flash(f"✅ Patient '{patient_name}' deleted successfully", "success")
    return redirect(url_for("patients.list_patients"))


# Export patients to Excel
@bp.route("/export", methods=["GET"])
@login_required
def export_patients():
    from flask_login import current_user

    # Permission check
    if not current_user.can_import_export():
        flash("⛔ You don't have permission to export. Contact your doctor.", "danger")
        return redirect(url_for("patients.list_patients"))

    from flask import send_file
    from io import BytesIO
    from flask_login import current_user

    # Get only current user's patients for export
    patients = (
        Patient.query.filter_by(doctor_id=current_user.id)
        .order_by(Patient.full_name.asc())
        .all()
    )

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Remove default sheet
    workbook.remove(workbook.active)

    # Sheet 1: Patient Summary
    summary_sheet = workbook.create_sheet("Patient Summary")
    summary_headers = [
        "#",
        "Full Name",
        "Phone",
        "Date of Birth",
        "Age",
        "Total Visits",
        "Last Visit",
        "Created Date",
    ]
    summary_sheet.append(summary_headers)

    # Style summary headers
    for cell in summary_sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(
            start_color="0070C0", end_color="0070C0", fill_type="solid"
        )
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center"
        )

    # Add patient summary data
    for idx, patient in enumerate(patients, start=1):
        # Calculate age
        age = ""
        if patient.date_of_birth:
            today = datetime.now().date()
            age = (
                today.year
                - patient.date_of_birth.year
                - (
                    (today.month, today.day)
                    < (patient.date_of_birth.month, patient.date_of_birth.day)
                )
            )

        # Get last visit date
        last_visit = ""
        if patient.visits:
            last_visit_obj = max(patient.visits, key=lambda v: v.visit_date)
            last_visit = last_visit_obj.visit_date.strftime("%Y-%m-%d")

        row = [
            idx,
            patient.full_name,
            patient.phone or "",
            patient.date_of_birth.strftime("%Y-%m-%d") if patient.date_of_birth else "",
            age,
            len(patient.visits),
            last_visit,
            patient.created_at.strftime("%Y-%m-%d"),
        ]
        summary_sheet.append(row)

    # Auto-adjust column widths for summary
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_sheet.column_dimensions[column_letter].width = adjusted_width

    # Sheet 2: Visits by Patient (Grouped)
    grouped_sheet = workbook.create_sheet("Visits by Patient")

    for patient in patients:
        if patient.visits:
            # Patient header
            grouped_sheet.append([])  # Empty row for spacing
            patient_header = grouped_sheet.cell(
                grouped_sheet.max_row + 1, 1, f"Patient: {patient.full_name}"
            )
            patient_header.font = openpyxl.styles.Font(
                bold=True, size=14, color="FFFFFF"
            )
            patient_header.fill = openpyxl.styles.PatternFill(
                start_color="0070C0", end_color="0070C0", fill_type="solid"
            )
            grouped_sheet.merge_cells(
                f"A{grouped_sheet.max_row}:H{grouped_sheet.max_row}"
            )

            # Patient info
            info_row = grouped_sheet.max_row + 1
            grouped_sheet.cell(info_row, 1, "Phone:")
            grouped_sheet.cell(info_row, 2, patient.phone or "—")
            grouped_sheet.cell(info_row, 3, "DOB:")
            grouped_sheet.cell(
                info_row,
                4,
                patient.date_of_birth.strftime("%Y-%m-%d")
                if patient.date_of_birth
                else "—",
            )

            # Visit headers
            visit_headers = [
                "Date",
                "Time",
                "Clinician",
                "Code",
                "Diagnosis",
                "Notes",
            ]
            grouped_sheet.append(visit_headers)
            header_row = grouped_sheet.max_row
            for col_num, cell in enumerate(grouped_sheet[header_row], start=1):
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )

            # Visit data
            for visit in sorted(
                patient.visits, key=lambda v: v.visit_date, reverse=True
            ):
                if visit.diagnoses:
                    for diagnosis in visit.diagnoses:
                        grouped_sheet.append(
                            [
                                visit.visit_date.strftime("%Y-%m-%d"),
                                visit.visit_date.strftime("%H:%M"),
                                visit.clinician or "",
                                diagnosis.code or "",
                                diagnosis.description,
                                visit.notes or "",
                            ]
                        )
                else:
                    grouped_sheet.append(
                        [
                            visit.visit_date.strftime("%Y-%m-%d"),
                            visit.visit_date.strftime("%H:%M"),
                            visit.clinician or "",
                            "",
                            "No diagnosis",
                            visit.notes or "",
                        ]
                    )

    # Auto-adjust column widths for grouped sheet
    for column in grouped_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        grouped_sheet.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Generate filename with current date
    filename = f"patients_export_{current_user.username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# Import patients from Excel
@bp.route("/import", methods=["GET", "POST"])
@login_required
def import_patients():
    from flask_login import current_user

    # Permission check
    if not current_user.can_import_export():
        flash("⛔ You don't have permission to import. Contact your doctor.", "danger")
        return redirect(url_for("patients.list_patients"))

    form = ImportPatientsForm()
    if form.validate_on_submit():
        file = form.file.data

        try:
            # Load the Excel file
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            imported_count = 0
            updated_count = 0
            skipped_count = 0
            duplicate_count = 0
            errors = []

            # Skip header row (assumes first row is header)
            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                try:
                    # Expected columns: Full Name, Phone, Date of Birth, Visit Date, Clinician, Notes, Diagnosis Code, Diagnosis Description
                    full_name = row[0] if len(row) > 0 else None
                    phone = row[1] if len(row) > 1 else None
                    dob_value = row[2] if len(row) > 2 else None
                    visit_date_value = row[3] if len(row) > 3 else None
                    clinician = row[4] if len(row) > 4 else None
                    notes = row[5] if len(row) > 5 else None
                    diagnosis_code = row[6] if len(row) > 6 else None
                    diagnosis_description = row[7] if len(row) > 7 else None

                    # Validate required fields
                    if not full_name or str(full_name).strip() == "":
                        skipped_count += 1
                        errors.append(f"Row {row_idx}: Missing full name")
                        continue

                    # Normalize data
                    full_name = str(full_name).strip()
                    phone = str(phone).strip() if phone else None

                    # Parse date of birth
                    date_of_birth = None
                    if dob_value:
                        if isinstance(dob_value, datetime):
                            date_of_birth = dob_value.date()
                        elif isinstance(dob_value, str):
                            try:
                                for fmt in [
                                    "%Y-%m-%d",
                                    "%m/%d/%Y",
                                    "%d/%m/%Y",
                                    "%Y/%m/%d",
                                ]:
                                    try:
                                        date_of_birth = datetime.strptime(
                                            dob_value, fmt
                                        ).date()
                                        break
                                    except ValueError:
                                        continue
                            except:
                                pass

                    # DUPLICATE CHECK: Use the same logic as new_patient
                    existing_patient = check_duplicate_patient(
                        full_name=full_name, phone=phone, doctor_id=current_user.id
                    )

                    # Create or update patient
                    if existing_patient:
                        # Patient exists - update their info if needed
                        patient = existing_patient

                        # Update DOB if provided and different
                        if date_of_birth and patient.date_of_birth != date_of_birth:
                            patient.date_of_birth = date_of_birth

                        # Update phone if provided and different
                        if phone and patient.phone != phone:
                            patient.phone = phone

                        # Check if this is truly a duplicate (same data) or an update
                        if (
                            patient.date_of_birth == date_of_birth
                            and patient.phone == phone
                        ):
                            duplicate_count += 1
                        else:
                            updated_count += 1
                    else:
                        # No duplicate - create new patient
                        patient = Patient(
                            full_name=full_name,
                            phone=phone,
                            date_of_birth=date_of_birth,
                            doctor_id=current_user.id,  # Always assign to current user
                        )
                        db.session.add(patient)
                        db.session.flush()  # Get patient.id for visit
                        imported_count += 1

                    # Create visit if visit date or diagnosis is provided
                    if visit_date_value or diagnosis_description:
                        # Parse visit date
                        visit_date = None
                        if visit_date_value:
                            if isinstance(visit_date_value, datetime):
                                visit_date = visit_date_value
                            elif isinstance(visit_date_value, str):
                                try:
                                    for fmt in [
                                        "%Y-%m-%d",
                                        "%m/%d/%Y",
                                        "%d/%m/%Y",
                                        "%Y/%m/%d",
                                        "%Y-%m-%d %H:%M:%S",
                                        "%m/%d/%Y %H:%M:%S",
                                    ]:
                                        try:
                                            visit_date = datetime.strptime(
                                                visit_date_value, fmt
                                            )
                                            break
                                        except ValueError:
                                            continue
                                except:
                                    pass

                        # If no visit date provided, use current datetime
                        if not visit_date:
                            visit_date = datetime.utcnow()

                        # Create visit
                        visit = Visit(
                            patient_id=patient.id,
                            visit_date=visit_date,
                            clinician=str(clinician).strip() if clinician else None,
                            notes=str(notes).strip() if notes else None,
                            created_by=current_user.id,  # Track who imported this
                        )
                        db.session.add(visit)
                        db.session.flush()  # Get visit.id for diagnosis

                        # Create diagnosis if description is provided
                        if diagnosis_description and str(diagnosis_description).strip():
                            diagnosis = Diagnosis(
                                visit_id=visit.id,
                                code=str(diagnosis_code).strip()
                                if diagnosis_code
                                else None,
                                description=str(diagnosis_description).strip(),
                            )
                            db.session.add(diagnosis)

                except Exception as e:
                    skipped_count += 1
                    errors.append(f"Row {row_idx}: {str(e)}")

            # Commit all changes at once
            db.session.commit()

            # Show results with detailed feedback
            if imported_count > 0:
                flash(
                    f"✅ Successfully imported {imported_count} new patient(s)",
                    "success",
                )
            if updated_count > 0:
                flash(
                    f"ℹ️ Updated {updated_count} existing patient(s) with new information",
                    "info",
                )
            if duplicate_count > 0:
                flash(
                    f"⚠️ Skipped {duplicate_count} duplicate(s) (patient already exists with same data)",
                    "warning",
                )
            if skipped_count > 0:
                flash(f"❌ Skipped {skipped_count} row(s) due to errors", "danger")

            # Show first few errors if any
            for error in errors[:5]:
                flash(error, "warning")
            if len(errors) > 5:
                flash(f"... and {len(errors) - 5} more errors", "warning")

            return redirect(url_for("patients.list_patients"))

        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error importing file: {str(e)}", "danger")

    return render_template("patients/import.html", form=form)
